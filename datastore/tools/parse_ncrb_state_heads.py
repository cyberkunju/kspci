#!/usr/bin/env python3
"""Parses NCRB Tables 1A.4 (IPC) and 1A.5 (SLL) into raw/ncrb_state_heads_raw.json.

Both sheets are laid out for print, not for machines: the header spans four merged
rows and the table is split into side-by-side panels that each restart with
SL / State/UT. Two things follow from that and neither is optional.

  1. Forward-filling merged header cells must reset at every panel boundary, or a
     label bleeds from one block into the next.
  2. The fill must reset hierarchically — a level only carries forward while its
     parent is unchanged. Filling each level independently produced keys such as
     "Dowry Deaths > Deaths due to other Negligence".

Keys are emitted hierarchically ("Burglary ... > B) Night") because NCRB reuses bare
sub-labels under different parents, so a leaf-only key would be ambiguous.

Only the incidence ("I") columns are read; the victim and rate columns are ignored.

Usage:  python3 datastore/tools/parse_ncrb_state_heads.py
Requires: openpyxl, and the two xlsx files in datastore/raw/ (see fetch-geo.sh).
"""

import json, os, re, sys
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
RAW = os.path.join(HERE, '..', 'raw')

def norm(s):
    return re.sub(r'\s+', ' ', str(s or '')).strip()

def num(v):
    if v is None: return None
    if isinstance(v, (int, float)): return float(v)
    s = str(v).replace(',', '').strip()
    return float(s) if re.fullmatch(r'-?\d+(\.\d+)?', s) else None

def parse(path, leaf_rows, measure_row, index_row, first_data_row, group_row):
    ws = load_workbook(path, data_only=True)[load_workbook(path, data_only=True).sheetnames[0]]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    ncol = max(len(r) for r in rows)
    def row(i):
        r = list(rows[i]); return r + [None] * (ncol - len(r))
    idx = [norm(v) for v in row(index_row)]
    grp = [norm(v) for v in row(group_row)]
    boundaries = {i for i in range(ncol)
                  if idx[i] in ('[1]', '[2]') or grp[i] in ('SL', 'State/UT')}
    levels = []
    for lr in leaf_rows:
        r = row(lr)
        parent = levels[-1] if levels else None
        filled, last = [], None
        for i in range(ncol):
            if i in boundaries: last = None
            if parent is not None and i > 0 and parent[i] != parent[i - 1]: last = None
            v = norm(r[i])
            if v: last = v
            filled.append(last)
        levels.append(filled)
    parts = [[] for _ in range(ncol)]
    for lv in levels:
        for i in range(ncol):
            if lv[i] and (not parts[i] or parts[i][-1] != lv[i]): parts[i].append(lv[i])
    leaf = [' > '.join(p) if p else None for p in parts]
    meas = [norm(v) for v in row(measure_row)]
    cols = {i: leaf[i] for i in range(ncol)
            if meas[i] == 'I' and leaf[i] and leaf[i] not in ('SL', 'State/UT')}
    out = {}
    for ri in range(first_data_row, len(rows)):
        r = row(ri)
        st = norm(r[1])
        if not st or st.endswith(':') or st.upper().startswith('TOTAL'): continue
        d = out.setdefault(st, {})
        for i, name in cols.items():
            v = num(r[i])
            if v is not None: d[name] = d.get(name, 0) + int(v)
    return out

def main():
    ipc_path = os.path.join(RAW, 'ncrb_ipc_state_heads_2022.xlsx')
    sll_path = os.path.join(RAW, 'ncrb_sll_state_heads_2022.xlsx')
    for p in (ipc_path, sll_path):
        if not os.path.exists(p):
            sys.exit(f'Missing {p}. Run ./datastore/fetch-geo.sh first.')
    ipc = parse(ipc_path, [2, 3, 4, 5], 6, 7, 9, 2)
    sll = parse(sll_path, [2, 3, 4, 5], 6, 7, 9, 2)
    merged = {}
    for src in (ipc, sll):
        for st, heads in src.items():
            merged.setdefault(st, {}).update(heads)
    out = os.path.join(RAW, 'ncrb_state_heads_raw.json')
    json.dump(merged, open(out, 'w'))
    heads = {k for h in merged.values() for k in h}
    print(f'{out} -> {len(merged)} states, {len(heads)} crime-head columns')

if __name__ == '__main__':
    main()
